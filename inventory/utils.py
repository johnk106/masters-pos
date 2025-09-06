# inventory/utils.py
import io
from django.http import FileResponse
from openpyxl import Workbook
# Temporarily disabled due to PIL issues
# from reportlab.lib.pagesizes import A4
# from reportlab.pdfgen import canvas
from django.db.models import Q, F

from .models import *

def get_products_queryset(search: str = None, low_stock: bool = False):
    qs = Product.objects.all().select_related('category', 'sub_category', 'units')
    if low_stock:
        qs = qs.filter(stock_entries__quantity__lte=F('stock_entries__quantity_alert'))
    if search:
        qs = qs.filter(
            Q(name__icontains=search) |
            Q(sku__icontains=search)
        )
    return qs.distinct()

def export_to_excel(qs, basename='products'):
    """
    Given a product queryset and a basename, returns a FileResponse
    for an in-memory .xlsx named e.g. 'products.xlsx' or 'low-stocks.xlsx'
    """
    wb = Workbook()
    ws = wb.active
    ws.append(['Name', 'SKU', 'Category', 'Stock Qty', 'Price'])
    for p in qs:
        stock = p.stock_entries.first()
        ws.append([
            p.name,
            p.sku,
            p.category.name if p.category else '',
            stock.quantity if stock else 0,
            float(stock.price) if stock else 0.0,
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"{basename}.xlsx"
    return FileResponse(buffer, as_attachment=True, filename=filename)

# Temporarily disabled due to PIL issues
# def export_to_pdf(qs, basename='products'):
#     """
#     Given a product queryset and a basename, returns a FileResponse
#     for an in-memory PDF named e.g. 'products.pdf' or 'low-stocks.pdf'
#     """
#     buffer = io.BytesIO()
##     p = canvas.Canvas(buffer, pagesize=A4)
#     width, height = A4
#     y = height - 50
#
#     # Header
#     p.setFont("Helvetica-Bold", 12)
#     p.drawString(40, y, f"{basename.replace('-', ' ').title()} Report")
#     p.setFont("Helvetica", 10)
#     y -= 30
#
#     # Column titles
#     p.drawString(40, y, "Name")
#     p.drawString(200, y, "SKU")
#     p.drawString(300, y, "Qty")
#     p.drawString(350, y, "Price")
#     y -= 20
#
#     # Rows
#     for prod in qs:
#         stock = prod.stock_entries.first()
#         qty = stock.quantity if stock else 0
#         price = float(stock.price) if stock else 0
#         p.drawString(40, y, prod.name[:30])
#         p.drawString(200, y, prod.sku)
#         p.drawString(300, y, str(qty))
#         p.drawString(350, y, f"{price:.2f}")
#         y -= 20
#         if y < 50:
#             p.showPage()
#             y = height - 50
#
#     p.save()
#     buffer.seek(0)
#     filename = f"{basename}.pdf"
#     return FileResponse(buffer, as_attachment=True, filename=filename)

def export_to_pdf(qs, basename='products'):
    """Temporarily disabled - PDF export unavailable"""
    from django.http import HttpResponse
    return HttpResponse("PDF export temporarily disabled", status=503)


def get_categories_queryset(search: str = None):
    qs = Category.objects.prefetch_related('sub_categories').order_by('name')
    if search:
        qs = qs.filter(
            Q(name__icontains=search) |
            Q(slug__icontains=search)
        )
    return qs.distinct()

def get_subcategories_queryset(search: str = None):
    qs = SubCategory.objects.select_related('category').order_by('name')
    if search:
        qs = qs.filter(
            Q(name__icontains=search) |
            Q(slug__icontains=search) |
            Q(category__name__icontains=search)
        )
    return qs.distinct()

# —————————————
# Excel exporters
# —————————————

def export_categories_excel(qs, basename='categories'):
    wb = Workbook()
    ws = wb.active
    ws.append(['Name', 'Slug', 'Status', 'Date Created', 'Sub-Count'])
    for cat in qs:
        ws.append([
            cat.name,
            cat.slug,
            'Active' if cat.status else 'Inactive',
            cat.date_created.strftime('%Y-%m-%d'),
            cat.sub_categories.count(),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=f'{basename}.xlsx')

def export_subcategories_excel(qs, basename='sub-categories'):
    wb = Workbook()
    ws = wb.active
    ws.append(['Name', 'Slug', 'Category', 'Status'])
    for sub in qs:
        ws.append([
            sub.name,
            sub.slug,
            sub.category.name,
            'Active' if sub.status else 'Inactive',
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=f'{basename}.xlsx')

# —————————————
# PDF exporters
# —————————————

def export_categories_pdf(qs, basename="export"):
    """Temporarily disabled - PDF export unavailable"""
    from django.http import HttpResponse
    return HttpResponse("PDF export temporarily disabled", status=503)

    for cat in qs:
        c.drawString(40, y, cat.name[:25])
        c.drawString(200, y, cat.slug[:15])
        c.drawString(320, y, 'Active' if cat.status else 'Inactive')
        c.drawString(420, y, str(cat.sub_categories.count()))
        y -= 20
        if y < 50:
            c.showPage(); y = h - 50

    c.save(); buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=f'{basename}.pdf')

def export_subcategories_pdf(qs, basename="export"):
    """Temporarily disabled - PDF export unavailable"""
    from django.http import HttpResponse
    return HttpResponse("PDF export temporarily disabled", status=503)

    for sub in qs:
        c.drawString(40, y, sub.name[:25])
        c.drawString(200, y, sub.slug[:15])
        c.drawString(320, y, sub.category.name[:15])
        c.drawString(440, y, 'Active' if sub.status else 'Inactive')
        y -= 20
        if y < 50:
            c.showPage(); y = h - 50

    c.save(); buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=f'{basename}.pdf')


def get_units_queryset(search: str = None):
    qs = Unit.objects.order_by('name')
    if search:
        qs = qs.filter(
            Q(name__icontains=search) |
            Q(short_name__icontains=search)
        )
    return qs

def get_variants_queryset(search: str = None):
    qs = Variant.objects.order_by('name')
    if search:
        qs = qs.filter(
            Q(name__icontains=search) |
            Q(values__icontains=search)
        )
    return qs

# —————————————
# Excel exporters
# —————————————

def export_units_excel(qs, basename='units'):
    wb = Workbook()
    ws = wb.active
    ws.append(['Name', 'Short Name', 'Status', 'Date Created'])
    for unit in qs:
        ws.append([
            unit.name,
            unit.short_name,
            'Active' if unit.status else 'Inactive',
            unit.date_created.strftime('%Y-%m-%d'),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=f'{basename}.xlsx')

def export_variants_excel(qs, basename='variants'):
    wb = Workbook()
    ws = wb.active
    ws.append(['Name', 'Values', 'Status', 'Date Created'])
    for var in qs:
        ws.append([
            var.name,
            var.values,
            'Active' if var.status else 'Inactive',
            var.date_created.strftime('%Y-%m-%d'),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=f'{basename}.xlsx')

# —————————————
# PDF exporters
# —————————————

def export_units_pdf(qs, basename="export"):
    """Temporarily disabled - PDF export unavailable"""
    from django.http import HttpResponse
    return HttpResponse("PDF export temporarily disabled", status=503)

    for unit in qs:
        c.drawString(40, y, unit.name[:25])
        c.drawString(220, y, unit.short_name[:10])
        c.drawString(340, y, 'Active' if unit.status else 'Inactive')
        c.drawString(440, y, unit.date_created.strftime('%Y-%m-%d'))
        y -= 20
        if y < 50:
            c.showPage(); y = h - 50

    c.save(); buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=f'{basename}.pdf')

def export_variants_pdf(qs, basename="export"):
    """Temporarily disabled - PDF export unavailable"""
    from django.http import HttpResponse
    return HttpResponse("PDF export temporarily disabled", status=503)

    for var in qs:
        c.drawString(40, y, var.name[:25])
        c.drawString(220, y, var.values[:25])
        c.drawString(440, y, 'Active' if var.status else 'Inactive')
        y -= 20
        if y < 50:
            c.showPage(); y = h - 50

    c.save(); buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=f'{basename}.pdf')